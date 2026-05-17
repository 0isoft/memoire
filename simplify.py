import re
import unicodedata
import numpy as np
import pandas as pd


# ============================================================
# CONFIG
# ============================================================

PROGRAM_PATH = "program_length.csv"
SIGNUP_PATH = "start.csv"
WEEKLY_PATH = "data2.csv"
ALGORITHM_FEEDBACK_PATH = "final_feedback.csv"
GARMIN_PATH = "all-sessions-cleaned.csv"

OUTPUT_CSV = "reduced_participant_table.csv"
OUTPUT_XLSX = "reduced_participant_table.xlsx"


# ============================================================
# BASIC HELPERS
# ============================================================

def normalize_text(x):
    """Lowercase, remove accents, normalize spaces."""
    if pd.isna(x):
        return ""

    x = str(x).strip().lower()
    x = unicodedata.normalize("NFKD", x)
    x = "".join(ch for ch in x if not unicodedata.combining(ch))
    x = x.replace("’", "'")
    x = re.sub(r"\s+", " ", x)
    return x.strip()


def normalize_column_name(x):
    """Normalize column names for robust matching."""
    x = normalize_text(x)
    x = re.sub(r"[\[\]\(\):?,;/_-]+", " ", x)
    x = re.sub(r"\s+", " ", x)
    return x.strip()


def clean_pid(series):
    """
    Convert PID formats:
    PP-6487 -> 6487
    PP 1550 -> 1550
    PP7056  -> 7056
    6487    -> 6487
    """
    return (
        series.astype(str)
        .str.strip()
        .str.replace("PP-", "", regex=False)
        .str.replace("pp-", "", regex=False)
        .str.extract(r"(\d+)", expand=False)
        .pipe(pd.to_numeric, errors="coerce")
    )


def to_numeric(series):
    return pd.to_numeric(series, errors="coerce")


def clean_group(series):
    return (
        series.astype(str)
        .str.strip()
        .str.upper()
        .replace({"": np.nan, "NAN": np.nan, "nan": np.nan})
    )


def clean_children_count(series):
    s = series.astype(str).str.strip().str.lower()
    s = (
        s.str.replace("4 ou plus", "4", regex=False)
         .str.replace("4+", "4", regex=False)
         .str.extract(r"(\d+)", expand=False)
    )
    return pd.to_numeric(s, errors="coerce")


def find_column(df, candidates, required=False):
    """
    Find a column by exact or partial normalized match.
    """
    normalized_cols = {col: normalize_column_name(col) for col in df.columns}

    for candidate in candidates:
        candidate_norm = normalize_column_name(candidate)

        # Exact match first
        for original_col, normalized_col in normalized_cols.items():
            if candidate_norm == normalized_col:
                return original_col

        # Partial match second
        for original_col, normalized_col in normalized_cols.items():
            if candidate_norm in normalized_col:
                return original_col

    if required:
        raise KeyError(
            f"Could not find required column matching {candidates}.\n"
            f"Available columns:\n{list(df.columns)}"
        )

    return None


# ============================================================
# 1. PROGRAM TABLE
# ============================================================

def load_program(path):
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()

    required = ["PID", "program_length", "group"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"program_length.csv is missing columns: {missing}")

    out = df[required].copy()
    out["PID"] = clean_pid(out["PID"])
    out["program_length"] = to_numeric(out["program_length"])
    out["group"] = clean_group(out["group"])

    out = out.dropna(subset=["PID", "program_length", "group"])
    out["PID"] = out["PID"].astype(int)
    out["program_length"] = out["program_length"].astype(int)

    out = out.drop_duplicates("PID", keep="first")
    return out.sort_values(["group", "PID"])


# ============================================================
# 2. SIGNUP / BASELINE TABLE
# ============================================================

def load_signup(path):
    df = pd.read_csv(path, low_memory=False)
    df.columns = df.columns.str.strip()

    pid_col = find_column(df, ["Colonne 1", "PID"], required=True)
    age_col = find_column(df, ["Votre âge", "votre age"], required=False)
    pp_col = find_column(df, ["à combien de semaines de post-partum êtes-vous"], required=False)
    children_col = find_column(df, ["Combien d'enfants avec vous"], required=False)
    status_col = find_column(df, ["Statut"], required=False)

    out = pd.DataFrame()
    out["PID"] = clean_pid(df[pid_col])

    out["age"] = to_numeric(df[age_col]) if age_col else np.nan
    out["postpartum_weeks"] = to_numeric(df[pp_col]) if pp_col else np.nan
    out["children_count"] = clean_children_count(df[children_col]) if children_col else np.nan

    if status_col:
        raw_status = df[status_col].astype(str).str.strip()
        out["raw_status"] = raw_status
        out["statut"] = raw_status.apply(
            lambda x: "valide" if normalize_text(x) == "valide" else "pas valide"
        )
    else:
        out["raw_status"] = np.nan
        out["statut"] = "pas valide"

    out = out.dropna(subset=["PID"])
    out["PID"] = out["PID"].astype(int)

    # One baseline row per participant
    out = out.drop_duplicates("PID", keep="first")

    return out[["PID", "age", "postpartum_weeks", "children_count", "statut"]]


# ============================================================
# 3. WEEKLY JDB / DATA2
# ============================================================

WEEKLY_SCORE_COLUMNS = {
    "seances": [
        "J'ai pu réaliser les séances prévues",
        "j ai pu realiser les seances prevues",
    ],
    "clarite": [
        "Les explications fournis étaient claires",
        "Les explications fournies étaient claires",
    ],
    "respect_consignes": [
        "J'ai respecté l'intensité/les consignes",
        "j ai respecte l intensite les consignes",
    ],
    "satisfaction_seance": [
        "Satisfaction globale de la séance",
    ],
    "confiance": [
        "Je me sens en confiance pour poursuivre le programme",
    ],
    "recommandation": [
        "Je recommanderais ces séances à une amie",
    ],
}


def load_weekly_normalized(path, df_program):
    df = pd.read_csv(path, low_memory=False)
    df.columns = df.columns.str.strip()

    pid_col = find_column(df, ["Votre PID", "PID"], required=True)
    week_col = find_column(
        df,
        ["A quelle semaine du programme êtes-vous arrivés", "semaine du programme"],
        required=True,
    )

    out = pd.DataFrame()
    out["PID"] = clean_pid(df[pid_col])
    out["week"] = to_numeric(df[week_col])

    for clean_name, candidates in WEEKLY_SCORE_COLUMNS.items():
        col = find_column(df, candidates, required=False)
        out[clean_name] = to_numeric(df[col]) if col else np.nan

    out = out.dropna(subset=["PID", "week"])
    out["PID"] = out["PID"].astype(int)

    # Merge program length
    out = out.merge(df_program[["PID", "program_length"]], on="PID", how="left")
    out = out.dropna(subset=["program_length"])

    # Keep only weeks inside the official program length
    out = out[
        (out["week"] >= 1)
        & (out["week"] <= out["program_length"])
    ].copy()

    weekly_cols = list(WEEKLY_SCORE_COLUMNS.keys())

    # ------------------------------------------------------------
    # If multiple forms exist for the same PID/week,
    # average them first.
    # This prevents duplicate forms from overweighting one week.
    # ------------------------------------------------------------
    pid_week = (
        out.groupby(["PID", "week"], as_index=False)[weekly_cols]
        .mean()
    )

    # Number of unique weeks with a form
    n_weeks_reported = (
        pid_week.groupby("PID")["week"]
        .nunique()
        .reset_index(name="n_formulaires_suivi")
    )

    # ------------------------------------------------------------
    # Correct satisfaction/adherence score:
    # average over submitted weeks only.
    #
    # This answers:
    # "When the participant filled the weekly form,
    # how positive was her answer on average?"
    # ------------------------------------------------------------
    means = (
        pid_week.groupby("PID")[weekly_cols]
        .mean()
        .reset_index()
    )

    # Add program length and number of submitted weeks
    means = means.merge(df_program[["PID", "program_length"]], on="PID", how="left")
    means = means.merge(n_weeks_reported, on="PID", how="left")

    # ------------------------------------------------------------
    # Follow-up completion:
    # number of submitted weeks / total expected program weeks.
    #
    # This is the variable that should penalize missing forms.
    # ------------------------------------------------------------
    means["completion_suivi"] = (
        means["n_formulaires_suivi"] / means["program_length"]
    )

    return means[
        [
            "PID",
            "seances",
            "clarite",
            "respect_consignes",
            "satisfaction_seance",
            "confiance",
            "recommandation",
            "n_formulaires_suivi",
            "completion_suivi",
        ]
    ]
# ============================================================
# 4. ALGORITHM PROCESS FEEDBACK
# ============================================================

ALGO_ISSUE_COLUMNS = {
    "probleme_contre_indications": ["Contre-indications"],
    "probleme_tests_fonctionnels": ["Tests fonctionnels"],
    "probleme_caracteristiques": ["Caractéristiques"],
    "probleme_criteres": ["Critères"],
    "probleme_symptomes": ["Symptômes"],
    "probleme_facilite_tests": ["Facilité des tests"],
    "probleme_choix_programme": ["Choix du programme"],
}

ALGO_SCORE_COLUMNS = {
    "utilisation_globale": ["Utilisation globale"],
    "deroulement": ["Déroulement"],
    "confiance_algorithme": ["Confiance"],
    "recommandation_algorithme": ["Recommandation"],
}


def encode_problem_cell(value):
    """
    Binary issue coding:
    - "Non" or empty -> 0
    - "Oui" or any free-text problem/comment -> 1
    """
    text = normalize_text(value)

    if text == "" or text == "nan":
        return 0

    if text == "non":
        return 0

    return 1


def load_algorithm_feedback(path):
    df = pd.read_csv(path, low_memory=False)
    df.columns = df.columns.str.strip()

    pid_col = find_column(df, ["PID"], required=True)

    out = pd.DataFrame()
    out["PID"] = clean_pid(df[pid_col])

    # Likert scores
    score_cols = []
    for clean_name, candidates in ALGO_SCORE_COLUMNS.items():
        col = find_column(df, candidates, required=True)
        out[clean_name] = to_numeric(df[col])
        score_cols.append(clean_name)

    out["satisfaction_generale_algorithme"] = out[score_cols].mean(axis=1, skipna=True)

    # Problem flags
    problem_cols = []
    for clean_name, candidates in ALGO_ISSUE_COLUMNS.items():
        col = find_column(df, candidates, required=True)
        out[clean_name] = df[col].apply(encode_problem_cell).astype(int)
        problem_cols.append(clean_name)

    out["nombre_problemes_algorithme"] = out[problem_cols].sum(axis=1)

    out = out.dropna(subset=["PID"])
    out["PID"] = out["PID"].astype(int)

    # If duplicate algorithm forms exist, average scores and sum/flag issues carefully.
    # For problems: max per problem keeps "reported at least once".
    agg = {
        "satisfaction_generale_algorithme": "mean",
        "utilisation_globale": "mean",
        "deroulement": "mean",
        "confiance_algorithme": "mean",
        "recommandation_algorithme": "mean",
    }

    for c in problem_cols:
        agg[c] = "max"

    out = out.groupby("PID", as_index=False).agg(agg)
    out["nombre_problemes_algorithme"] = out[problem_cols].sum(axis=1)

    return out[
        [
            "PID",
            "satisfaction_generale_algorithme",
            "nombre_problemes_algorithme",
        ]
    ]


# ============================================================
# 5. NOLIO / GARMIN
# ============================================================

def load_garmin_summary(path):
    df = pd.read_csv(path, low_memory=False)
    df.columns = df.columns.str.strip()

    # Some files use "id", others "PID"
    if "PID" not in df.columns and "id" in df.columns:
        df["PID"] = df["id"]

    if "PID" not in df.columns:
        raise KeyError("Garmin file must contain either PID or id column.")

    df["PID"] = clean_pid(df["PID"])
    df = df.dropna(subset=["PID"])
    df["PID"] = df["PID"].astype(int)

    # Use only clean included sessions if the column exists
    if "include_in_garmin_analysis" in df.columns:
        include = (
            df["include_in_garmin_analysis"]
            .astype(str)
            .str.strip()
            .str.lower()
            .isin(["true", "1", "yes", "oui"])
        )
        df = df[include].copy()

    if len(df) == 0:
        return pd.DataFrame(columns=[
            "PID",
            "nombre_sessions_nolio",
            "reached_30min_ever",
        ])

    if "max_continuous_run_min" not in df.columns:
        raise KeyError("Garmin file must contain max_continuous_run_min column.")

    df["max_continuous_run_min"] = to_numeric(df["max_continuous_run_min"])

    summary = (
        df.groupby("PID")
        .agg(
            nombre_sessions_nolio=("PID", "count"),
            garmin_max_run=("max_continuous_run_min", "max"),
        )
        .reset_index()
    )

    summary["reached_30min_ever"] = summary["garmin_max_run"] >= 30

    return summary[
        [
            "PID",
            "nombre_sessions_nolio",
            "reached_30min_ever",
        ]
    ]


# ============================================================
# 6. BUILD FINAL REDUCED TABLE
# ============================================================

def build_reduced_table():
    program = load_program(PROGRAM_PATH)
    signup = load_signup(SIGNUP_PATH)
    weekly = load_weekly_normalized(WEEKLY_PATH, program)
    algo = load_algorithm_feedback(ALGORITHM_FEEDBACK_PATH)
    garmin = load_garmin_summary(GARMIN_PATH)

    # Start from official participant list
    final = program.copy()

    final = final.merge(signup, on="PID", how="left")
    final = final.merge(weekly, on="PID", how="left")
    final = final.merge(algo, on="PID", how="left")
    final = final.merge(garmin, on="PID", how="left")

    # Fill missing indicator values
    final["statut"] = final["statut"].fillna("pas valide")
    final["nombre_sessions_nolio"] = final["nombre_sessions_nolio"].fillna(0).astype(int)
    final["reached_30min_ever"] = final["reached_30min_ever"].fillna(False).astype(bool)
    final["n_formulaires_suivi"] = final["n_formulaires_suivi"].fillna(0).astype(int)
    final["completion_suivi"] = final["completion_suivi"].fillna(0)

    # Final column order: only what you requested
    final = final[
    [
        "PID",
        "group",
        "age",
        "postpartum_weeks",
        "children_count",
        "statut",

        # Adhesion et Satisfaction — JDB / weekly averages
        "seances",
        "clarite",
        "respect_consignes",
        "satisfaction_seance",
        "confiance",
        "recommandation",
        "n_formulaires_suivi",
        "completion_suivi",

        # Avis sur l'algorithme
        "satisfaction_generale_algorithme",
        "nombre_problemes_algorithme",

        # Nolio
        "nombre_sessions_nolio",
        "reached_30min_ever",
    ]
].sort_values("PID")

    return final


# ============================================================
# 7. EXPORT
# ============================================================

def export_table(df):
    df.to_csv(OUTPUT_CSV, index=False)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Table")

        ws = writer.book["Table"]

        # Freeze header
        ws.freeze_panes = "A2"

        # Simple column widths
        widths = {
            "A": 10,  # PID
            "B": 8,   # age
            "C": 18,  # postpartum
            "D": 16,  # children
            "E": 14,  # statut
            "F": 12,
            "G": 12,
            "H": 20,
            "I": 22,
            "J": 14,
            "K": 18,
            "L": 20,
            "M": 18,
            "N": 32,
            "O": 28,
            "P": 22,
            "Q": 18,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Number formats
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

    print(f"Saved CSV:  {OUTPUT_CSV}")
    print(f"Saved XLSX: {OUTPUT_XLSX}")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    table = build_reduced_table()
    export_table(table)

    print("\nPreview:")
    print(table.head(10).to_string(index=False))